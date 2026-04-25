"""
Email & Phone Enrichment Tool
=============================
Reads any CSV that contains a website/URL column and scrapes each site
for contact email addresses AND phone numbers using two sequential passes:

  Pass 1 — Fast HTTP GET  (homepage + configured contact paths, no browser)
  Pass 2 — Playwright     (headless Chromium fallback for JS-rendered sites)

Results are written to an Excel workbook (+ CSV backup) with a Run Stats sheet.
Progress is checkpointed so any interrupted run can be resumed.
Auto-saves every N sites AND every 60 seconds in a background thread.

Column detection is fully automatic — no need to rename CSV headers.
Input file detection is automatic — no need to rename your file.

Usage:
  python enricher.py                          # auto-detects CSV in current dir
  python enricher.py --input my_leads.csv
  python enricher.py --input leads.csv --config my_config.yaml
  python enricher.py --fresh                  # clear checkpoint and restart
  python enricher.py --output results.xlsx

Runtime controls (while running):
  P  — pause / resume
  R  — resume
  Q  — quit (saves progress first)
  S  — print current status
  Windows: single-key (no Enter needed)
  Mac/Linux: type the letter then press Enter

Requirements:
  pip install requests playwright pyyaml openpyxl tqdm
  python -m playwright install chromium
"""

import os
import re
import csv
import json
import sys
import time
import random
import socket
import shutil
import select
import threading
import argparse
import platform
import urllib.robotparser
from datetime import datetime, date
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import requests
import urllib3
import yaml

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Optional tqdm — falls back gracefully if not installed
try:
    from tqdm import tqdm as _TqdmClass
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

    class _TqdmClass:           # type: ignore[no-redef]
        """Minimal no-op shim used when tqdm is not installed."""
        def __init__(self, *a, **kw) -> None:
            self.total = kw.get("total", 0)
            self.n = 0
        def update(self, n: int = 1) -> None:   self.n += n
        def set_postfix(self, **kw) -> None:    pass
        def write(self, s: str) -> None:        print(s, flush=True)
        def close(self) -> None:                pass
        def __enter__(self):                    return self
        def __exit__(self, *a) -> None:         pass


# ================================================================
# CONFIG LOADER
# ================================================================

DEFAULT_CONFIG: dict = {
    # ── Input / Output ───────────────────────────────────────────
    # Leave blank to auto-detect a CSV in the current directory.
    "input_file":            "",
    "output_file":           "",
    "checkpoint_file":       "enrich_checkpoint.json",
    "command_file":          "command.txt",
    "output_format":         "xlsx",   # "xlsx" or "csv"

    # ── Timing & Performance ─────────────────────────────────────
    "http_timeout":          [4, 6],
    "playwright_timeout":    8000,
    "browser_restart_every": 150,
    "stop_at":               "23:00",
    "autosave_interval":     60,       # background save every N seconds

    # ── Rate Limiting ────────────────────────────────────────────
    "rate_limit": {
        "min_seconds": 0.1,
        "max_seconds": 0.5,
    },

    # ── User-Agent Rotation ──────────────────────────────────────
    "user_agents": [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.0; rv:109.0) Gecko/20100101 Firefox/120.0",
    ],

    # ── Scraping Behaviour ───────────────────────────────────────
    "contact_paths": [
        "/contact", "/contact-us", "/about", "/about-us",
    ],
    "locale": "en-US",

    # ── Column Names ─────────────────────────────────────────────
    # Leave blank ("") for auto-detection from your CSV headers.
    # These are the OUTPUT column names written to the results file.
    "columns": {
        "company_name": "",       # auto-detect
        "website":      "Website",
        "email":        "Email",
        "phone":        "Phone",
        "category":     "",       # auto-detect (optional)
    },

    # ── Email Filtering ──────────────────────────────────────────
    "skip_email_keywords": [
        "noreply", "no-reply", "donotreply", "privacy", "dataprotection",
        "data-protection", "gdpr", "unsubscribe", "postmaster", "webmaster",
        "bounce", "complaints", "legal", "abuse", "spam", "newsletter",
    ],
    "generic_email_keywords": [
        "info", "admin", "hello", "contact", "enquiries", "enquiry", "office",
        "mail", "email", "team", "support", "help", "sales", "lettings",
        "letting", "property", "management", "manager", "reception",
        "accounts", "finance", "general", "service", "post",
    ],
    "junk_email_domains": [
        "sentry.io", "wixpress.com", "example.com", "schema.org", "w3.org",
        "googleapis.com", "cloudflare.com", "jquery.com",
        "domain.com", "email.com", "test.com", "mailinator.com",  # placeholder domains
    ],

    # ── Cookie Banner Dismissal ──────────────────────────────────
    "cookie_selectors": [
        'button:has-text("Accept all")',    'button:has-text("Accept All")',
        'button:has-text("Accept cookies")', 'button:has-text("Accept")',
        'button:has-text("I Accept")',      'button:has-text("Allow all")',
        'button:has-text("OK")',            'button:has-text("Got it")',
        '[id*="accept"]',                   '[aria-label*="Accept"]',
    ],
}


def load_config(config_path: Optional[str] = None) -> dict:
    """
    Load configuration from a YAML file, merged on top of defaults.

    Nested dicts (e.g. rate_limit, columns) are shallow-merged so partial
    YAML overrides don't wipe unmentioned sub-keys.
    """
    cfg = {k: (v.copy() if isinstance(v, dict) else v) for k, v in DEFAULT_CONFIG.items()}
    if config_path and os.path.exists(config_path):
        with open(config_path, encoding="utf-8") as f:
            user_cfg = yaml.safe_load(f) or {}
        for key, val in user_cfg.items():
            if isinstance(val, dict) and isinstance(cfg.get(key), dict):
                cfg[key].update(val)
            else:
                cfg[key] = val
    return cfg


# ================================================================
# LOGGING
# ================================================================

_start_time: float = time.time()
_active_bar: Optional[object] = None   # set to a tqdm instance during passes


def elapsed() -> str:
    """Return elapsed time since script start formatted as [Xm YYs]."""
    s = int(time.time() - _start_time)
    return f"[{s // 60}m{s % 60:02d}s]"


def log(msg: str, kind: str = "info") -> None:
    """Print a timestamped, icon-prefixed status line.
    Routes through tqdm.write() when a progress bar is active to avoid glitches."""
    icons = {"good": "✅", "warn": "⚠ ", "error": "❌", "info": "  ", "dim": "  "}
    text = f"{elapsed():>9} {icons.get(kind, '  ')} {msg}"
    if _active_bar is not None:
        _active_bar.write(text)   # type: ignore[attr-defined]
    else:
        print(text, flush=True)


# ================================================================
# AUTO-SAVER  (background thread)
# ================================================================

class AutoSaver:
    """
    Background daemon thread that persists results every `interval` seconds.
    This ensures data is never lost even between the regular per-site saves.

    Usage:
        saver = AutoSaver(found, out_file, cfg, stats, interval=60)
        # … run the pass …
        saver.stop()
    """

    def __init__(
        self,
        found:    dict,
        out_file: str,
        cfg:      dict,
        stats:    dict,
        interval: int = 60,
    ) -> None:
        self._found    = found
        self._out_file = out_file
        self._cfg      = cfg
        self._stats    = stats
        self._interval = max(1, interval)
        self._stopped  = False
        t = threading.Thread(target=self._run, daemon=True)
        t.start()

    def _run(self) -> None:
        ticks = 0
        while not self._stopped:
            time.sleep(1)
            ticks += 1
            if ticks >= self._interval and not self._stopped:
                ticks = 0
                try:
                    save_output(self._found, self._out_file, self._cfg, self._stats)
                except Exception:
                    pass  # never crash the background thread

    def stop(self) -> None:
        """Signal the thread to stop (it is a daemon so it won't block exit)."""
        self._stopped = True


# ================================================================
# CROSS-PLATFORM CONTROLS
# ================================================================

class State:
    """Shared mutable run state — accessed from both main thread and ControlListener."""
    paused: bool = False
    stop:   bool = False


class ControlListener:
    """
    Listens for interactive keyboard commands in a background daemon thread.

    Platform behaviour:
      Windows  — msvcrt single-key detection (no Enter required).
      Mac/Linux — select-based stdin line reading (type letter + Enter).

    Supported commands:
      P — toggle pause/resume
      R — resume (if paused)
      Q — quit (saves before exit)
      S — print current status
    """

    def __init__(self, state: State, ctx: dict) -> None:
        self._state = state
        self._ctx   = ctx
        t = threading.Thread(target=self._listen, daemon=True)
        t.start()

    def _handle(self, key: str) -> None:
        key = key.strip().upper()
        if not key:
            return
        key = key[0]
        s   = self._state

        if key == "P":
            s.paused = not s.paused
            if s.paused:
                log("PAUSED — press P or R to resume", "warn");  _beep("stop")
            else:
                log("RESUMED", "good");                           _beep("resume")
        elif key == "R" and s.paused:
            s.paused = False
            log("RESUMED", "good");                               _beep("resume")
        elif key == "Q":
            s.stop = True
            log("QUIT — saving and exiting …", "warn");           _beep("stop")
        elif key == "S":
            log(f"status → found:{self._ctx.get('found', 0)} | done:{self._ctx.get('done', 0)}")

    def _listen(self) -> None:
        if platform.system() == "Windows":
            import msvcrt
            while True:
                if msvcrt.kbhit():
                    try:
                        key = msvcrt.getch().decode(errors="ignore")
                        while msvcrt.kbhit():
                            msvcrt.getch()
                        self._handle(key)
                    except Exception:
                        pass
                time.sleep(0.05)
        else:
            # Unix / macOS: block-read lines from stdin
            while True:
                try:
                    ready, _, _ = select.select([sys.stdin], [], [], 0.2)
                    if ready:
                        line = sys.stdin.readline()
                        if line:
                            self._handle(line.strip())
                except Exception:
                    time.sleep(0.1)


def _beep(kind: str = "error") -> None:
    """
    Emit an audio alert.
    Windows: winsound frequency sequences.
    All other platforms: console bell character (silent if terminal mutes it).
    """
    try:
        if platform.system() == "Windows":
            import winsound
            B = winsound.Beep
            sequences = {
                "start":  [(500, 100), (700, 100), (900, 200)],
                "resume": [(600, 150), (900, 250)],
                "done":   [(600, 100), (800, 100), (1000, 100), (1200, 300)],
                "stop":   [(900, 200), (600, 400)],
            }
            for freq, dur in sequences.get(kind, [(350, 120)]):
                B(freq, dur)
        else:
            print("\a", end="", flush=True)
    except Exception:
        pass


def check_cmd_file(state: State, cmd_file: str, checkpoint_file: str) -> None:
    """
    Read a single command from the command file and clear it.
    Valid file contents: pause | resume | stop | fresh
    """
    if not os.path.exists(cmd_file):
        return
    try:
        with open(cmd_file, encoding="utf-8") as fh:
            cmd = fh.read().strip().lower()
        if not cmd:
            return
        with open(cmd_file, "w", encoding="utf-8") as fh:
            fh.write("")

        if cmd == "pause":
            state.paused = True;  log("PAUSED (cmd file)", "warn");   _beep("stop")
        elif cmd in ("resume", "r"):
            state.paused = False; log("RESUMED (cmd file)", "good");   _beep("resume")
        elif cmd in ("stop", "q"):
            state.stop = True;    log("STOP — saving …", "warn");      _beep("stop")
        elif cmd == "fresh":
            if os.path.exists(checkpoint_file):
                os.remove(checkpoint_file)
            log("Checkpoint cleared — restart the script to begin fresh", "warn")
    except Exception:
        pass


def wait_if_paused(state: State, ctx: dict, cmd_file: str, checkpoint_file: str) -> None:
    """Block execution until the paused flag is cleared or a stop is requested."""
    while state.paused and not state.stop:
        check_cmd_file(state, cmd_file, checkpoint_file)
        time.sleep(0.3)


def should_stop(state: State, stop_at: str) -> bool:
    """Return True if a quit was requested, or if the wall-clock stop time has passed."""
    if state.stop:
        return True
    if not stop_at:
        return False
    try:
        return datetime.now().strftime("%H:%M") >= stop_at
    except Exception:
        return False


# ================================================================
# SYSTEM CHECKS
# ================================================================

def has_internet() -> bool:
    """Test internet connectivity by attempting a TCP connection to a public DNS server."""
    try:
        socket.setdefaulttimeout(3)
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect(("8.8.8.8", 53))
        s.close()
        return True
    except Exception:
        return False


def wait_for_internet(state: State) -> None:
    """Auto-pause and poll every 30 s until internet connectivity is restored."""
    if has_internet():
        return
    log("Internet unreachable — auto-pausing, retrying every 30 s", "warn")
    _beep("stop")
    state.paused = True
    attempt = 0
    while not has_internet() and not state.stop:
        attempt += 1
        if attempt % 2 == 0:
            log(f"Still waiting for internet … ({attempt * 30}s elapsed)", "warn")
        time.sleep(30)
    if has_internet():
        state.paused = False
        log("Internet restored — resuming", "good")
        _beep("resume")


def check_disk(min_mb: int = 500) -> bool:
    """Warn and return False if free disk space falls below min_mb megabytes."""
    free_mb = shutil.disk_usage(".").free // (1024 * 1024)
    if free_mb < min_mb:
        log(f"Low disk space: {free_mb} MB free (minimum {min_mb} MB) — pausing", "warn")
        _beep("stop")
        return False
    return True


# ================================================================
# EMAIL UTILITIES
# ================================================================

def extract_emails_raw(html: str) -> List[str]:
    """
    Extract plaintext email addresses from HTML using a permissive regex.
    Filters out asset-path false positives and addresses longer than 80 chars.
    """
    raw = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", html)
    result = []
    for e in raw:
        e = e.lower().strip().strip('.,;"\'')
        if not re.match(r"^[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}$", e):
            continue
        if any(ext in e for ext in [".png", ".jpg", ".jpeg", ".svg", ".gif", ".css", ".js"]):
            continue
        if len(e) > 80:
            continue
        result.append(e)
    return list(set(result))


def decode_cloudflare_email(encoded: str) -> str:
    """
    Decode a Cloudflare email-protection hex string.
    Cloudflare XORs each byte against the first byte (the key), then hex-encodes.
    """
    try:
        enc = bytes.fromhex(encoded)
        key = enc[0]
        return "".join(chr(b ^ key) for b in enc[1:])
    except Exception:
        return ""


def extract_emails_full(html: str) -> List[str]:
    """
    Extract all email addresses from HTML, including those obfuscated by
    Cloudflare's email-protection system (cdn-cgi and data-cfemail variants).
    """
    emails = extract_emails_raw(html)
    for pattern in (
        r"/cdn-cgi/l/email-protection#([a-f0-9]+)",
        r'data-cfemail="([a-f0-9]+)"',
    ):
        for m in re.finditer(pattern, html):
            decoded = decode_cloudflare_email(m.group(1))
            if "@" in decoded:
                emails.append(decoded.lower().strip())
    return list(set(emails))


def extract_phones(html: str) -> List[str]:
    """
    Extract phone numbers from HTML.

    Strategy (priority order):
      1. tel: href attributes — highest confidence, very low false-positive rate.
      2. Regex pattern matching — broader sweep filtered by digit count (7–15).

    Returns a deduplicated list of raw phone strings, most reliable first.
    """
    seen_digits: Set[str] = set()
    phones: List[str] = []

    def _add(raw: str) -> None:
        raw = raw.strip()
        digits = re.sub(r"\D", "", raw)
        if 7 <= len(digits) <= 15 and digits not in seen_digits:
            seen_digits.add(digits)
            phones.append(raw)

    # 1. tel: href links — best source
    for m in re.finditer(r'href=["\']tel:([^"\']{4,25})["\']', html, re.I):
        _add(m.group(1).replace("%20", " ").replace("+", "+"))

    # 2. Regex fallback when no tel: links are present
    if not phones:
        patterns = [
            # International: +44 20 7123 4567, +1-800-555-5555
            r"\+\d{1,3}[\s\-\.]?\(?\d{1,4}\)?[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}",
            # Bracketed area code: (020) 7123 4567
            r"\(\d{2,5}\)[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}",
            # Plain runs with separators: 020 7123 4567, 555-555-5555
            r"\b\d{3,5}[\s\-\.]\d{3,4}[\s\-\.]\d{3,4}\b",
        ]
        for pat in patterns:
            for m in re.finditer(pat, html):
                _add(m.group(0))

    return phones


def score_email(email: str, cfg: dict) -> int:
    """
    Score an email address by contact quality. Lower = better.

    Score  Meaning
    -----  -------
      1    Personal name  (e.g. john.smith@company.com)  — most valuable
      2    High-priority generic  (info@, hello@, contact@)
      3    Other generic  (support@, accounts@, sales@)
    999    Junk / skip-list  — filtered out entirely
    """
    if not email or "@" not in email:
        return 999
    local  = email.lower().split("@")[0]
    domain = email.lower().split("@")[1]

    skip_kws     = set(cfg.get("skip_email_keywords",   []))
    generic_kws  = set(cfg.get("generic_email_keywords", []))
    junk_domains = set(cfg.get("junk_email_domains",     []))

    if any(k in local  for k in skip_kws):     return 999
    if domain in junk_domains:                  return 999
    if not any(k in local for k in generic_kws): return 1   # personal name
    if local in {"info", "hello", "contact", "enquiries", "enquiry"}: return 2
    return 3


def best_email(emails: List[str], cfg: dict) -> str:
    """Return the single highest-quality email from a list, or '' if none qualify."""
    scored = [(e.lower().strip(), score_email(e, cfg)) for e in emails if e and "@" in e]
    valid  = [(e, s) for e, s in scored if s < 999]
    if not valid:
        return ""
    return min(valid, key=lambda x: x[1])[0]


def best_phone(phones: List[str]) -> str:
    """
    Return the single highest-confidence phone number from a list.

    Preference order:
      1. International format starting with +  (e.g. +44 20 7123 4567)
      2. Any other number with a valid digit count (7–15 digits)

    extract_phones() already surfaces tel: href numbers first, so index 0
    is usually correct — but this scorer adds an explicit quality gate so
    obviously bad strings (too few digits, truncated) are never returned.
    """
    if not phones:
        return ""

    def _score(p: str) -> int:
        digits = re.sub(r"\D", "", p)
        if not 7 <= len(digits) <= 15:
            return 999
        if p.strip().startswith("+"):
            return 1
        return 2

    scored = [(p, _score(p)) for p in phones]
    valid  = [(p, s) for p, s in scored if s < 999]
    if not valid:
        return phones[0]   # fallback — return raw rather than nothing
    return min(valid, key=lambda x: x[1])[0]


# ================================================================
# I/O UTILITIES
# ================================================================

def get_output_path(cfg: dict) -> str:
    """Resolve the output file path, auto-generating a timestamped name if unset."""
    if cfg.get("output_file"):
        return cfg["output_file"]
    ext = "xlsx" if cfg.get("output_format", "xlsx") == "xlsx" else "csv"
    return f"found_contacts_{date.today().strftime('%Y%m%d')}.{ext}"


def _detect_column(headers: List[str], *keywords: str) -> Optional[str]:
    """Return the first header containing any keyword (case-insensitive). None if not found."""
    for h in headers:
        h_lower = h.lower()
        if any(kw in h_lower for kw in keywords):
            return h
    return None


def find_input_file() -> Optional[str]:
    """
    Auto-detect the input CSV file from the current working directory.

    - Exactly one CSV → use it automatically.
    - Multiple CSVs   → prompt the user to choose.
    - None found      → return None (caller will error out with guidance).
    """
    csv_files = sorted(Path(".").glob("*.csv"))

    if not csv_files:
        return None

    if len(csv_files) == 1:
        return str(csv_files[0])

    # Multiple files — ask the user
    print("\nMultiple CSV files found in this directory. Please choose one:")
    for i, f in enumerate(csv_files, 1):
        print(f"  {i}. {f.name}")
    print()
    while True:
        try:
            raw = input("Enter number (or full path to another file): ").strip()
            if raw.isdigit():
                idx = int(raw) - 1
                if 0 <= idx < len(csv_files):
                    return str(csv_files[idx])
            elif os.path.exists(raw):
                return raw
            print("  Invalid choice — try again.")
        except (EOFError, KeyboardInterrupt):
            return None


def load_input(cfg: dict) -> List[dict]:
    """
    Load and validate the input CSV with fully automatic column detection.

    Detection order for each column type:
      Website (REQUIRED):   website · url · domain · site · web · link · homepage
      Company name (opt):   company · name · organisation · organization · business
                            firm · client · account · brand · title
                            → falls back to the first column if nothing matches
      Category (opt):       category · type · sector · industry · segment · group · vertical
      Pre-existing phone:   phone · tel · mobile · cell · number · contact number

    Raises FileNotFoundError or ValueError on critical problems.
    Returns list of target dicts: {key, name, website, phone, category}.
    """
    input_file = cfg["input_file"]

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    with open(input_file, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        raise ValueError("Input file is empty.")

    headers: List[str] = list(rows[0].keys())
    cols = cfg.get("columns", {})

    # ── Website column (required) ────────────────────────────────
    col_web = cols.get("website") or ""
    if not col_web or col_web not in headers:
        col_web = _detect_column(
            headers,
            "website", "url", "domain", "site", "web", "link", "homepage",
        )
    if not col_web:
        raise ValueError(
            f"Cannot find a website/URL column in your CSV.\n"
            f"Columns found: {headers}\n"
            f"Please add a column named 'Website', 'URL', or 'Domain'."
        )

    # ── Company name column (optional — falls back to first column) ──
    col_name = cols.get("company_name") or ""
    if not col_name or col_name not in headers:
        col_name = _detect_column(
            headers,
            "company", "name", "organisation", "organization",
            "business", "firm", "client", "account", "brand", "title",
        )
    if not col_name:
        col_name = headers[0]
        log(f"No company-name column detected — using first column: '{col_name}'", "warn")

    # ── Category column (optional) ───────────────────────────────
    col_cat = cols.get("category") or ""
    if not col_cat or col_cat not in headers:
        col_cat = _detect_column(
            headers,
            "category", "type", "sector", "industry",
            "segment", "group", "vertical",
        )
    # col_cat may be None — that's fine, category is optional

    # ── Pre-existing phone column (optional) ─────────────────────
    col_phone_in = _detect_column(
        headers,
        "phone", "tel", "mobile", "cell", "number", "contact number",
    )

    log(
        f"Columns → name='{col_name}'  website='{col_web}'"
        + (f"  category='{col_cat}'" if col_cat else "")
        + (f"  phone_in='{col_phone_in}'" if col_phone_in else "")
    )

    return [
        {
            "key":      row[col_name].strip().lower(),
            "name":     row[col_name].strip(),
            "website":  row[col_web].strip(),
            # Carry through any pre-existing phone so it's preserved in output
            "phone":    row.get(col_phone_in, "").strip() if col_phone_in else "",
            "category": row.get(col_cat, "").strip()      if col_cat      else "",
        }
        for row in rows
        if row.get(col_web, "").strip()   # website is the only truly required field
    ]


def _csv_path(path: str) -> str:
    """Return the CSV companion path for a given output path."""
    return path.replace(".xlsx", ".csv") if path.endswith(".xlsx") else path


def load_existing_output(path: str, cfg: dict) -> dict:
    """
    Load previously found contacts from an existing output CSV (for resume support).
    A row is loaded if it has a non-empty email OR phone.
    """
    csv_p = _csv_path(path)
    if not os.path.exists(csv_p):
        return {}

    cols      = cfg.get("columns", {})
    col_name  = cols.get("company_name") or "Company Name"
    col_web   = cols.get("website")      or "Website"
    col_email = cols.get("email")        or "Email"
    col_phone = cols.get("phone")        or "Phone"
    col_cat   = cols.get("category")     or "Category"

    found: dict = {}
    with open(csv_p, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = row.get(col_name, "").strip().lower()
            if key and (row.get(col_email, "").strip() or row.get(col_phone, "").strip()):
                found[key] = {
                    "name":     row.get(col_name,  ""),
                    "website":  row.get(col_web,   ""),
                    "email":    row.get(col_email, ""),
                    "phone":    row.get(col_phone, ""),
                    "category": row.get(col_cat,   ""),
                }
    return found


def save_output_csv(found: dict, path: str, cfg: dict) -> None:
    """Write found contacts to a UTF-8 CSV file (always produced as a resume backup)."""
    if not found:
        return
    cols      = cfg.get("columns", {})
    col_name  = cols.get("company_name") or "Company Name"
    col_web   = cols.get("website")      or "Website"
    col_email = cols.get("email")        or "Email"
    col_phone = cols.get("phone")        or "Phone"
    col_cat   = cols.get("category")     or "Category"
    csv_p     = _csv_path(path)

    with open(csv_p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[col_name, col_web, col_email, col_phone, col_cat])
        w.writeheader()
        for v in found.values():
            w.writerow({
                col_name:  v.get("name",     ""),
                col_web:   v.get("website",  ""),
                col_email: v.get("email",    ""),
                col_phone: v.get("phone",    ""),
                col_cat:   v.get("category", ""),
            })


def save_output_xlsx(found: dict, path: str, cfg: dict, stats: dict) -> None:
    """
    Write results to an Excel workbook with two sheets:

      Sheet 1 — Results   : one row per company with found contact data
      Sheet 2 — Run Stats : summary metrics for the enrichment run

    Falls back to CSV-only if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("openpyxl not installed — falling back to CSV. Run: pip install openpyxl", "warn")
        return

    if not found:
        return

    cols      = cfg.get("columns", {})
    col_name  = cols.get("company_name") or "Company Name"
    col_web   = cols.get("website")      or "Website"
    col_email = cols.get("email")        or "Email"
    col_phone = cols.get("phone")        or "Phone"
    col_cat   = cols.get("category")     or "Category"

    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Results"

    HDR_FILL = PatternFill("solid", fgColor="2E4057")   # dark navy
    HDR_FONT = Font(color="FFFFFF", bold=True)
    headers  = [col_name, col_web, col_email, col_phone, col_cat]

    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for ri, v in enumerate(found.values(), 2):
        ws1.cell(row=ri, column=1, value=v.get("name",     ""))
        ws1.cell(row=ri, column=2, value=v.get("website",  ""))
        ws1.cell(row=ri, column=3, value=v.get("email",    ""))
        ws1.cell(row=ri, column=4, value=v.get("phone",    ""))
        ws1.cell(row=ri, column=5, value=v.get("category", ""))

    for ci, width in enumerate([40, 45, 40, 25, 30], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = width

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Run Stats ────────────────────────────────────────
    ws2 = wb.create_sheet("Run Stats")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    STAT_FILL = PatternFill("solid", fgColor="4A7C59")  # muted green
    STAT_FONT = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(["Metric", "Value"], 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = STAT_FILL
        cell.font = STAT_FONT

    total   = stats.get("total", 0)
    n_email = sum(1 for v in found.values() if v.get("email"))
    n_phone = sum(1 for v in found.values() if v.get("phone"))
    n_any   = len(found)
    pct_e   = f"{round(n_email / total * 100)}%" if total else "N/A"
    pct_p   = f"{round(n_phone / total * 100)}%" if total else "N/A"
    pct_any = f"{round(n_any   / total * 100)}%" if total else "N/A"

    stat_rows = [
        ("Run Timestamp",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Input File",         stats.get("input_file", "")),
        ("Companies Input",    total),
        ("Contacts Found",     n_any),
        ("  — Emails Found",   n_email),
        ("  — Phones Found",   n_phone),
        ("Email Success Rate", pct_e),
        ("Phone Success Rate", pct_p),
        ("Any Contact Rate",   pct_any),
        ("Still Missing",      total - n_any),
        ("Pass 1 Found",       stats.get("pass1_found", 0)),
        ("Pass 2 Found",       stats.get("pass2_found", 0)),
        ("Time Elapsed",       stats.get("elapsed",     "")),
    ]
    for ri, (metric, value) in enumerate(stat_rows, 2):
        ws2.cell(row=ri, column=1, value=metric)
        ws2.cell(row=ri, column=2, value=value)

    xlsx_path = path if path.endswith(".xlsx") else path.replace(".csv", ".xlsx")
    wb.save(xlsx_path)


def save_output(found: dict, path: str, cfg: dict, stats: Optional[dict] = None) -> None:
    """
    Persist current results.
    Always writes a CSV (used for resume logic).
    Also writes XLSX with a Stats sheet when output_format is 'xlsx'.
    """
    save_output_csv(found, path, cfg)
    if cfg.get("output_format", "xlsx") == "xlsx":
        save_output_xlsx(found, path, cfg, stats or {})


def save_checkpoint(done: Set[str], found: dict, checkpoint_file: str) -> None:
    """Atomically persist the current done-set and found-map to a JSON checkpoint."""
    with open(checkpoint_file, "w", encoding="utf-8") as f:
        json.dump({"done": list(done), "found": found}, f)


def load_checkpoint(checkpoint_file: str) -> Tuple[Set[str], dict]:
    """Load a previously saved checkpoint. Returns (done_set, found_dict)."""
    if not os.path.exists(checkpoint_file):
        return set(), {}
    try:
        content = open(checkpoint_file, encoding="utf-8").read().strip()
        if not content:
            return set(), {}
        data = json.loads(content)
        return set(data.get("done", [])), data.get("found", {})
    except Exception:
        return set(), {}


# ================================================================
# NETWORK — helpers
# ================================================================

_robots_cache: Dict[str, bool] = {}


def _can_fetch(url: str) -> bool:
    """
    Return True if the site's robots.txt permits scraping *url*.

    Two performance guarantees:
      - Short timeout (3 s) via requests — urllib's default has no timeout
        and can stall a run for 30+ seconds per site.
      - Per-domain caching — robots.txt is fetched at most ONCE per site
        even when both passes visit the same domain.

    Fails open (returns True) on any network or parse error so a missing
    or unreachable robots.txt never silently blocks the run.
    """
    try:
        from urllib.parse import urlparse
        parsed     = urlparse(url if url.startswith("http") else "https://" + url)
        domain_key = f"{parsed.scheme}://{parsed.netloc}"

        if domain_key in _robots_cache:
            return _robots_cache[domain_key]

        robots_url = f"{domain_key}/robots.txt"
        try:
            resp    = requests.get(
                robots_url, timeout=3, verify=False,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            content = resp.text if resp.status_code == 200 else ""
        except Exception:
            _robots_cache[domain_key] = True
            return True

        rp = urllib.robotparser.RobotFileParser()
        rp.set_url(robots_url)
        rp.parse(content.splitlines())
        allowed = rp.can_fetch("*", url)
        _robots_cache[domain_key] = allowed
        return allowed
    except Exception:
        return True   # fail open

def random_ua(cfg: dict) -> str:
    """Pick a random User-Agent string from the configured pool."""
    pool = cfg.get("user_agents", [])
    return random.choice(pool) if pool else "Mozilla/5.0"


def _rate_limit(cfg: dict) -> None:
    """Sleep for a random duration within the configured [min, max] range."""
    rl = cfg.get("rate_limit", {})
    lo = float(rl.get("min_seconds", 0.0))
    hi = float(rl.get("max_seconds", 0.3))
    if hi > 0:
        time.sleep(random.uniform(lo, hi))


# ================================================================
# PASS 1 — requests GET
# ================================================================

def _fetch_worker(url: str, ua: str, timeout: tuple, result: list) -> None:
    """Daemon-thread worker: fetch a single URL and append HTML to result list."""
    try:
        headers = {
            "User-Agent":      ua,
            "Accept":          "text/html,*/*",
            "Accept-Language": "en-US,en;q=0.9",
        }
        r = requests.get(url, headers=headers, timeout=timeout,
                         verify=False, allow_redirects=True)
        if r.status_code < 400:
            result.append(r.text)
    except Exception:
        pass


def fetch_page(url: str, cfg: dict, wall_clock_limit: int = 10) -> Optional[str]:
    """
    Fetch a URL with a hard wall-clock timeout enforced via a daemon thread.
    The daemon thread is abandoned (not killed) once the timeout expires;
    because it is a daemon it will not prevent process exit.
    """
    result: list = []
    t = threading.Thread(
        target=_fetch_worker,
        args=(url, random_ua(cfg), tuple(cfg.get("http_timeout", [4, 6])), result),
        daemon=True,
    )
    t.start()
    try:
        t.join(timeout=wall_clock_limit)
    except KeyboardInterrupt:
        return None
    return result[0] if result else None


def enrich_one_http(target: dict, cfg: dict) -> Tuple[str, str]:
    """
    Pass 1: attempt to find a contact email AND phone via plain HTTP GET requests.

    Checks robots.txt before fetching. Visits the homepage first, then iterates
    through contact_paths from cfg. Stops early once a high-quality (score ≤ 2)
    email is found. Rate limiting and UA rotation are applied on every request.

    Returns (best_email, best_phone).  Either may be an empty string.
    """
    base          = target["website"].rstrip("/")
    contact_paths = cfg.get("contact_paths", ["/contact", "/about"])
    emails: List[str] = []
    phones: List[str] = []

    if not _can_fetch(base):
        log(f"robots.txt disallows: {base} — skipping", "warn")
        return "", ""

    # Homepage
    html = fetch_page(base, cfg)
    if html:
        emails.extend(extract_emails_full(html))
        phones.extend(extract_phones(html))
    _rate_limit(cfg)

    # Early exit if an excellent email is already found
    if any(score_email(e, cfg) <= 2 for e in emails):
        return best_email(emails, cfg), best_phone(phones)

    # Contact / about pages
    for path in contact_paths:
        html = fetch_page(base + path, cfg)
        if html:
            found_emails = extract_emails_full(html)
            found_phones = extract_phones(html)
            emails.extend(found_emails)
            phones.extend(found_phones)
            if any(score_email(e, cfg) <= 2 for e in found_emails):
                break
        _rate_limit(cfg)

    return best_email(emails, cfg), best_phone(phones)


# ================================================================
# PASS 2 — Playwright headless browser
# ================================================================

def _launch_browser(p, cfg: dict):
    """
    Launch a headless Chromium instance with media and tracking routes blocked.
    Retries up to 3 times before raising. Returns (browser, page) tuple.
    """
    ua     = random_ua(cfg)
    locale = cfg.get("locale", "en-US")

    BLOCKED = [
        "**/*.{png,jpg,jpeg,gif,svg,webp,ico,woff,woff2,ttf,eot,mp4,mp3}",
        "**/google-analytics**",
        "**/googletagmanager**",
        "**/doubleclick**",
    ]

    for attempt in range(3):
        try:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage",
                    "--disable-extensions", "--mute-audio",
                    "--disable-background-networking",
                ],
            )
            ctx = browser.new_context(
                user_agent=ua,
                locale=locale,
                ignore_https_errors=True,
            )
            for pattern in BLOCKED:
                ctx.route(pattern, lambda route: route.abort())
            return browser, ctx.new_page()
        except Exception as e:
            log(f"Browser launch attempt {attempt + 1}/3 failed: {e}", "warn")
            time.sleep(3)

    raise RuntimeError("Chromium failed to launch after 3 attempts.")


def _dismiss_cookie_banner(page, cfg: dict) -> None:
    """Silently attempt to click any recognised cookie consent button."""
    for sel in cfg.get("cookie_selectors", []):
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=1000):
                btn.click(timeout=800)
                time.sleep(0.15)
                return
        except Exception:
            pass


def enrich_one_browser(page, target: dict, cfg: dict) -> Tuple[str, str]:
    """
    Pass 2: find a contact email AND phone using a Playwright-rendered browser.

    Checks robots.txt before fetching. Handles JS-rendered pages that return no
    extractable HTML to plain requests. Visits homepage + first two contact_paths
    from cfg. Dismisses cookie banners on the homepage visit.

    Returns (best_email, best_phone).  Either may be an empty string.
    """
    base          = target["website"].rstrip("/")
    contact_paths = cfg.get("contact_paths", ["/contact", "/about"])
    pw_timeout    = cfg.get("playwright_timeout", 8000)
    emails: List[str] = []
    phones: List[str] = []

    if not _can_fetch(base):
        log(f"robots.txt disallows: {base} — skipping", "warn")
        return "", ""

    urls = [base] + [base + p for p in contact_paths[:2]]

    for i, url in enumerate(urls):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=pw_timeout)
            if i == 0:
                _dismiss_cookie_banner(page, cfg)
            content = page.content()
            emails.extend(extract_emails_full(content))
            phones.extend(extract_phones(content))
            if any(score_email(e, cfg) <= 2 for e in emails):
                break
        except Exception:
            continue
        _rate_limit(cfg)

    return best_email(emails, cfg), best_phone(phones)


# ================================================================
# PASS RUNNERS
# ================================================================

def run_pass1(
    targets:  List[dict],
    done:     Set[str],
    found:    dict,
    out_file: str,
    state:    State,
    ctx:      dict,
    cfg:      dict,
) -> List[dict]:
    """
    Execute Pass 1: fast HTTP enrichment for all targets not yet in found.

    Saves progress every 10 sites.
    Returns a list of targets that yielded no contact data and should be
    retried in Pass 2 using Playwright.
    """
    global _active_bar

    todo     = [t for t in targets if t["key"] not in found]
    stop_at  = cfg.get("stop_at", "")
    ckpt     = cfg["checkpoint_file"]
    cmd_file = cfg["command_file"]

    log(f"Pass 1 — {len(todo)} sites → requests GET "
        f"(homepage + {len(cfg.get('contact_paths', []))} contact paths)")

    if not todo:
        log("Pass 1: nothing to process")
        return []

    needs_pw:   List[dict] = []
    pass1_found = 0
    fail_streak = 0

    bar = _TqdmClass(
        total=len(todo),
        desc="  Pass 1 (HTTP)   ",
        unit="site",
        dynamic_ncols=True,
        colour="cyan" if TQDM_AVAILABLE else None,
    )
    _active_bar = bar

    try:
        for count, target in enumerate(todo, 1):
            check_cmd_file(state, cmd_file, ckpt)
            if should_stop(state, stop_at): break
            wait_if_paused(state, ctx, cmd_file, ckpt)
            if should_stop(state, stop_at): break

            email, phone = enrich_one_http(target, cfg)
            done.add(target["key"])
            ctx["done"] = len(done)

            if email or phone:
                found[target["key"]] = {
                    "name":     target["name"],
                    "website":  target["website"],
                    "email":    email,
                    # Prefer freshly scraped phone; fall back to pre-existing
                    "phone":    phone or target.get("phone", ""),
                    "category": target["category"],
                }
                pass1_found += 1
                ctx["found"] = pass1_found
                fail_streak  = 0
            else:
                # Carry over any pre-existing phone data from the input CSV
                if target.get("phone"):
                    found[target["key"]] = {
                        "name":     target["name"],
                        "website":  target["website"],
                        "email":    "",
                        "phone":    target["phone"],
                        "category": target["category"],
                    }
                    pass1_found += 1
                    ctx["found"] = pass1_found
                needs_pw.append(target)
                fail_streak += 1
                if fail_streak % 3 == 0:
                    wait_for_internet(state)
                    if should_stop(state, stop_at): break
                    fail_streak = 0

            # Save every 10 sites
            if count % 10 == 0:
                save_checkpoint(done, found, ckpt)
                save_output(found, out_file, cfg)
                check_disk()

            pct   = round(pass1_found / count * 100)
            rem   = len(todo) - count
            eta   = int(rem * (time.time() - _start_time) / max(count, 1) / 60)
            eta_s = f"~{eta // 60}h{eta % 60:02d}m" if eta >= 60 else f"~{eta}m"
            bar.set_postfix(found=pass1_found, hit=f"{pct}%", eta=eta_s)
            bar.update(1)

    finally:
        _active_bar = None
        bar.close()

    print()
    save_checkpoint(done, found, ckpt)
    save_output(found, out_file, cfg)
    log(f"Pass 1 done — {pass1_found} contacts found, "
        f"{len(needs_pw)} sites queued for Playwright", "good")
    return needs_pw


def run_pass2(
    needs_pw: List[dict],
    done:     Set[str],
    found:    dict,
    out_file: str,
    state:    State,
    ctx:      dict,
    cfg:      dict,
    stats:    dict,
) -> None:
    """
    Execute Pass 2: Playwright-based enrichment for JS-heavy sites.

    The browser is restarted every browser_restart_every sites to prevent
    memory accumulation during large runs. Progress is saved every 10 sites.
    """
    global _active_bar

    todo          = [t for t in needs_pw if t["key"] not in found]
    stop_at       = cfg.get("stop_at", "")
    ckpt          = cfg["checkpoint_file"]
    cmd_file      = cfg["command_file"]
    restart_every = cfg.get("browser_restart_every", 150)

    log(f"Pass 2 — {len(todo)} sites → Playwright headless browser")

    if not todo:
        log("Pass 2: nothing to process")
        return

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log(
            "playwright not installed — "
            "run: pip install playwright && python -m playwright install chromium",
            "error",
        )
        return

    pass2_found = 0
    pw_count    = 0

    bar = _TqdmClass(
        total=len(todo),
        desc="  Pass 2 (Browser)",
        unit="site",
        dynamic_ncols=True,
        colour="green" if TQDM_AVAILABLE else None,
    )
    _active_bar = bar

    try:
        with sync_playwright() as p:
            browser, page = _launch_browser(p, cfg)

            for count, target in enumerate(todo, 1):
                check_cmd_file(state, cmd_file, ckpt)
                if should_stop(state, stop_at): break
                wait_if_paused(state, ctx, cmd_file, ckpt)
                if should_stop(state, stop_at): break

                # Periodic browser restart prevents memory accumulation
                if pw_count > 0 and pw_count % restart_every == 0:
                    log(f"Restarting browser after {pw_count} sites …", "dim")
                    try: browser.close()
                    except Exception: pass
                    time.sleep(2)
                    browser, page = _launch_browser(p, cfg)

                email, phone = enrich_one_browser(page, target, cfg)
                done.add(target["key"])
                pw_count += 1
                ctx["done"] = len(done)

                if email or phone:
                    found[target["key"]] = {
                        "name":     target["name"],
                        "website":  target["website"],
                        "email":    email,
                        "phone":    phone or target.get("phone", ""),
                        "category": target["category"],
                    }
                    pass2_found += 1
                    ctx["found"]         = pass2_found
                    stats["pass2_found"] = pass2_found

                # Save every 10 sites
                if count % 10 == 0:
                    if not check_disk(): break
                    save_checkpoint(done, found, ckpt)
                    save_output(found, out_file, cfg, stats)
                    wait_for_internet(state)
                    if should_stop(state, stop_at): break

                rem   = len(todo) - count
                eta   = int(rem * 3 / 60)
                eta_s = f"~{eta // 60}h{eta % 60:02d}m" if eta >= 60 else f"~{eta}m"
                pct   = round(pass2_found / count * 100)
                bar.set_postfix(found=pass2_found, hit=f"{pct}%", eta=eta_s)
                bar.update(1)
                time.sleep(0.1)

            try: browser.close()
            except Exception: pass

    finally:
        _active_bar = None
        bar.close()

    print()
    save_checkpoint(done, found, ckpt)
    save_output(found, out_file, cfg, stats)
    log(f"Pass 2 done — {pass2_found} additional contacts found via Playwright", "good")


# ================================================================
# MAIN
# ================================================================

def parse_args() -> argparse.Namespace:
    """Define and parse CLI arguments."""
    parser = argparse.ArgumentParser(
        prog="enricher",
        description=(
            "Email & Phone Enrichment Tool — "
            "scrape contact emails and phone numbers from company websites."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python enricher.py                          # auto-detect CSV in current directory
  python enricher.py --input companies.csv
  python enricher.py --input leads.csv --config my_config.yaml
  python enricher.py --fresh                  # ignore checkpoint, start over
  python enricher.py --output results.xlsx    # override output path
        """,
    )
    parser.add_argument("--input",  "-i", help="Path to input CSV (overrides config and auto-detection)")
    parser.add_argument("--output", "-o", help="Path to output file (overrides config)")
    parser.add_argument("--config", "-c", default="config.yaml",
                        help="Path to YAML config file (default: config.yaml)")
    parser.add_argument("--fresh",  "-f", action="store_true",
                        help="Clear existing checkpoint and start from scratch")
    return parser.parse_args()


def _print_banner() -> None:
    os_name = platform.system()
    print()
    print("╔══════════════════════════════════════════════════╗")
    print("║   Email & Phone Enrichment Tool                  ║")
    print("║   Pass 1: HTTP  →  Pass 2: Playwright fallback   ║")
    print("╚══════════════════════════════════════════════════╝")
    if os_name == "Windows":
        print("  Controls: P=pause  R=resume  Q=quit  S=status")
    else:
        print("  Controls: type P / R / Q / S  then  Enter")
    if not TQDM_AVAILABLE:
        print("  Tip: pip install tqdm  for a richer progress bar")
    print()


def _print_summary(
    targets:  List[dict],
    found:    dict,
    out_file: str,
    stats:    dict,
    partial:  bool = False,
) -> None:
    total   = len(targets)
    n_email = sum(1 for v in found.values() if v.get("email"))
    n_phone = sum(1 for v in found.values() if v.get("phone"))
    n_any   = len(found)
    print()
    print("╔══════════════════════════════════════════════════╗")
    log(f"  {'PARTIAL — re-run to continue' if partial else 'COMPLETE'}")
    log(f"  Companies input  : {total}")
    log(f"  Contacts found   : {n_any}  ({round(n_any / total * 100) if total else 0}%)")
    log(f"    — Emails       : {n_email}  ({round(n_email / total * 100) if total else 0}%)")
    log(f"    — Phones       : {n_phone}  ({round(n_phone / total * 100) if total else 0}%)")
    log(f"  Still missing    : {total - n_any}")
    log(f"  Pass 1 found     : {stats.get('pass1_found', 0)}")
    log(f"  Pass 2 found     : {stats.get('pass2_found', 0)}")
    log(f"  Time elapsed     : {elapsed()}")
    log(f"  Output           : {os.path.abspath(out_file)}", "good")
    print("╚══════════════════════════════════════════════════╝")
    print()


def main() -> None:
    """
    Orchestrate the two-pass email & phone enrichment pipeline.

      1. Load config (YAML + CLI overrides)
      2. Auto-detect or validate input CSV
      3. Auto-detect columns (website, name, category, phone)
      4. Pass 1 — fast HTTP requests with background auto-save
      5. Pass 2 — Playwright fallback for JS-heavy sites with background auto-save
      6. Save Excel + CSV output with run statistics
    """
    global _start_time
    _start_time = time.time()

    args = parse_args()
    cfg  = load_config(args.config)

    # CLI arguments override config file values
    if args.input:  cfg["input_file"]  = args.input
    if args.output: cfg["output_file"] = args.output

    # Work relative to the script's own directory
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    _print_banner()

    # ── Auto-detect input file ────────────────────────────────────
    if not cfg.get("input_file"):
        detected = find_input_file()
        if not detected:
            log("No input CSV found in current directory.", "error")
            log("Usage: python enricher.py --input path/to/file.csv", "info")
            return
        cfg["input_file"] = detected
        log(f"Auto-detected input: {detected}", "good")

    # ── State and statistics ──────────────────────────────────────
    state = State()
    ctx:   dict = {"found": 0, "done": 0}
    stats: dict = {
        "pass1_found": 0,
        "pass2_found": 0,
        "total":       0,
        "elapsed":     "",
        "input_file":  cfg["input_file"],
    }

    ckpt = cfg["checkpoint_file"]

    if args.fresh and os.path.exists(ckpt):
        os.remove(ckpt)
        log("Checkpoint cleared — starting fresh", "warn")

    log(f"Config : {args.config}")
    log(f"Input  : {cfg['input_file']}")
    out_file = get_output_path(cfg)
    log(f"Output : {out_file}  [{cfg.get('output_format', 'xlsx').upper()}]")
    print()

    # Start keyboard listener (background daemon thread)
    ControlListener(state, ctx)

    try:
        targets = load_input(cfg)
    except (FileNotFoundError, ValueError) as e:
        log(str(e), "error")
        return

    stats["total"] = len(targets)
    log(f"Loaded {len(targets)} rows from CSV")

    if not targets:
        log("Nothing to process.", "warn")
        return

    # Category distribution (optional)
    cats = Counter(t["category"] for t in targets if t.get("category"))
    if cats:
        log("Category breakdown:")
        for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
            log(f"  {cat:<42} {n}", "dim")
    print()

    # Resume from checkpoint
    _, found = load_checkpoint(ckpt)
    found.update(load_existing_output(out_file, cfg))
    done: Set[str] = set()

    if found:
        log(f"Resuming — {len(found)} contacts already in cache", "good")
        _beep("resume")
    else:
        log("Fresh start", "good")
        _beep("start")

    ctx["done"] = len(found)

    autosave_interval = cfg.get("autosave_interval", 60)

    # ── Pass 1 ────────────────────────────────────────────────────
    auto_saver1 = AutoSaver(found, out_file, cfg, stats, interval=autosave_interval)
    needs_pw = run_pass1(targets, done, found, out_file, state, ctx, cfg)
    stats["pass1_found"] = len(found)
    auto_saver1.stop()

    if should_stop(state, cfg.get("stop_at", "")):
        save_checkpoint(done, found, ckpt)
        save_output(found, out_file, cfg, stats)
        _print_summary(targets, found, out_file, stats, partial=True)
        return

    print()

    # ── Pass 2 ────────────────────────────────────────────────────
    auto_saver2 = AutoSaver(found, out_file, cfg, stats, interval=autosave_interval)
    run_pass2(needs_pw, done, found, out_file, state, ctx, cfg, stats)
    auto_saver2.stop()

    stats["elapsed"] = elapsed()
    save_checkpoint(done, found, ckpt)
    save_output(found, out_file, cfg, stats)

    all_done = not state.stop
    if all_done and os.path.exists(ckpt):
        os.remove(ckpt)
    _beep("done" if all_done else "stop")

    _print_summary(targets, found, out_file, stats, partial=not all_done)


if __name__ == "__main__":
    main()
