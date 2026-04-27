"""
core.storage — Checkpoint persistence and Excel / CSV output.

Checkpoint writes are **atomic**: data is written to a ``.tmp`` file
first, then ``os.replace()`` swaps it into place in a single OS-level
operation.  This prevents a corrupt or empty checkpoint if the process
is killed mid-write.

Public API
----------
save_checkpoint(done, found, path)              — atomic JSON checkpoint write.
load_checkpoint(path)                           — load checkpoint → (done, found).
get_output_path(cfg)                            — resolve / auto-generate output path.
load_existing_output(path, cfg)                 — load contacts from a previous CSV.
save_output(found, path, cfg, stats)            — write CSV + XLSX results.
"""

from __future__ import annotations

import csv
import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional, Set, Tuple

from core._log import log


# ── Path helpers ───────────────────────────────────────────────────────────────

def get_output_path(cfg: dict) -> str:
    """
    Resolve the output file path.

    If ``cfg["output_file"]`` is set, use it verbatim.
    Otherwise generate a timestamped name: ``found_contacts_YYYYMMDD.<ext>``.
    """
    if cfg.get("output_file"):
        return cfg["output_file"]
    ext = "xlsx" if cfg.get("output_format", "xlsx") == "xlsx" else "csv"
    return f"found_contacts_{date.today().strftime('%Y%m%d')}.{ext}"


def _csv_path(path: str) -> str:
    """Return the companion CSV path for a given output path."""
    return path.replace(".xlsx", ".csv") if path.endswith(".xlsx") else path


# ── Checkpoint ─────────────────────────────────────────────────────────────────

def save_checkpoint(done: Set[str], found: dict, checkpoint_file: str) -> None:
    """
    Atomically persist the current done-set and found-map to a JSON checkpoint.

    Write strategy (crash-safe):
      1. Serialise to ``<checkpoint_file>.tmp``
      2. ``os.replace()`` — atomic on all POSIX systems and on Windows (NTFS)
         since Python 3.3.  No partial-write corruption possible.

    Parameters
    ----------
    done            : Set of company-key strings already processed.
    found           : Dict of found contact records keyed by company-key.
    checkpoint_file : Target file path (e.g. ``"enrich_checkpoint.json"``).
    """
    tmp = checkpoint_file + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({"done": list(done), "found": found}, f)
    os.replace(tmp, checkpoint_file)


def load_checkpoint(checkpoint_file: str) -> Tuple[Set[str], dict]:
    """
    Load a previously saved checkpoint.

    Returns ``(set(), {})`` if the file does not exist or is corrupt —
    safe to call unconditionally at startup.

    Returns
    -------
    ``(done_set, found_dict)``
    """
    if not os.path.exists(checkpoint_file):
        return set(), {}
    try:
        content = Path(checkpoint_file).read_text(encoding="utf-8").strip()
        if not content:
            return set(), {}
        data = json.loads(content)
        return set(data.get("done", [])), data.get("found", {})
    except Exception:
        return set(), {}


# ── CSV helpers ────────────────────────────────────────────────────────────────

def _col_names(cfg: dict) -> Tuple[str, str, str, str, str]:
    """Return resolved (company_name, website, email, phone, category) column names."""
    cols = cfg.get("columns", {})
    return (
        cols.get("company_name") or "Company Name",
        cols.get("website")      or "Website",
        cols.get("email")        or "Email",
        cols.get("phone")        or "Phone",
        cols.get("category")     or "Category",
    )


def load_existing_output(path: str, cfg: dict) -> dict:
    """
    Load previously found contacts from an existing output CSV.

    Used during resume: merges prior results into ``found`` before the run
    starts so already-enriched sites are never re-scraped.

    A row is loaded only if it has a non-empty email **or** phone value.

    Returns
    -------
    Dict keyed by lowercased company name — same structure as ``found``.
    """
    csv_p = _csv_path(path)
    if not os.path.exists(csv_p):
        return {}

    col_name, col_web, col_email, col_phone, col_cat = _col_names(cfg)
    found: dict = {}

    with open(csv_p, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = row.get(col_name, "").strip().lower()
            if key and (row.get(col_email, "").strip() or row.get(col_phone, "").strip()):
                found[key] = {
                    "name":     row.get(col_name,  ""),
                    "website":  row.get(col_web,   ""),
                    "email":    row.get(col_email, ""),
                    "phone":    row.get(col_phone, ""),
                    "category": row.get(col_cat,   ""),
                }
    return found


def _save_csv(found: dict, path: str, cfg: dict) -> None:
    """Write found contacts to a UTF-8 CSV file (always produced as a resume backup)."""
    if not found:
        return
    col_name, col_web, col_email, col_phone, col_cat = _col_names(cfg)
    csv_p = _csv_path(path)

    with open(csv_p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[col_name, col_web, col_email, col_phone, col_cat],
        )
        w.writeheader()
        for v in found.values():
            w.writerow({
                col_name:  v.get("name",     ""),
                col_web:   v.get("website",  ""),
                col_email: v.get("email",    ""),
                col_phone: v.get("phone",    ""),
                col_cat:   v.get("category", ""),
            })


# ── Excel output ───────────────────────────────────────────────────────────────

def _save_xlsx(found: dict, path: str, cfg: dict, stats: dict) -> None:
    """
    Write results to an Excel workbook with two sheets.

    Sheet 1 — Results
        One row per company with: Company Name, Website, Email, Phone, Category.
        Styled with dark-navy header row, frozen top row.

    Sheet 2 — Run Stats
        Summary metrics: totals, hit rates, pass breakdown, elapsed time.
        Styled with muted-green header row.

    Falls back to CSV-only if ``openpyxl`` is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("openpyxl not installed — falling back to CSV. Run: pip install openpyxl", "warn")
        return

    if not found:
        return

    col_name, col_web, col_email, col_phone, col_cat = _col_names(cfg)
    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Results"

    HDR_FILL = PatternFill("solid", fgColor="2E4057")   # dark navy
    HDR_FONT = Font(color="FFFFFF", bold=True)
    sheet_headers = [col_name, col_web, col_email, col_phone, col_cat]

    for ci, h in enumerate(sheet_headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for ri, v in enumerate(found.values(), 2):
        ws1.cell(row=ri, column=1, value=v.get("name",     ""))
        ws1.cell(row=ri, column=2, value=v.get("website",  ""))
        ws1.cell(row=ri, column=3, value=v.get("email",    ""))
        ws1.cell(row=ri, column=4, value=v.get("phone",    ""))
        ws1.cell(row=ri, column=5, value=v.get("category", ""))

    for ci, width in enumerate([40, 45, 40, 25, 30], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = width

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Run Stats ────────────────────────────────────────
    ws2 = wb.create_sheet("Run Stats")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    STAT_FILL = PatternFill("solid", fgColor="4A7C59")  # muted green
    STAT_FONT = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(["Metric", "Value"], 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = STAT_FILL
        cell.font = STAT_FONT

    total   = stats.get("total", 0)
    n_email = sum(1 for v in found.values() if v.get("email"))
    n_phone = sum(1 for v in found.values() if v.get("phone"))
    n_any   = len(found)
    pct_e   = f"{round(n_email / total * 100)}%" if total else "N/A"
    pct_p   = f"{round(n_phone / total * 100)}%" if total else "N/A"
    pct_any = f"{round(n_any   / total * 100)}%" if total else "N/A"

    stat_rows = [
        ("Run Timestamp",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Input File",         stats.get("input_file", "")),
        ("Companies Input",    total),
        ("Contacts Found",     n_any),
        ("  — Emails Found",   n_email),
        ("  — Phones Found",   n_phone),
        ("Email Success Rate", pct_e),
        ("Phone Success Rate", pct_p),
        ("Any Contact Rate",   pct_any),
        ("Still Missing",      total - n_any),
        ("Pass 1 Found",       stats.get("pass1_found", 0)),
        ("Pass 2 Found",       stats.get("pass2_found", 0)),
        ("Time Elapsed",       stats.get("elapsed",     "")),
    ]
    for ri, (metric, value) in enumerate(stat_rows, 2):
        ws2.cell(row=ri, column=1, value=metric)
        ws2.cell(row=ri, column=2, value=value)

    xlsx_path = path if path.endswith(".xlsx") else path.replace(".csv", ".xlsx")
    wb.save(xlsx_path)


# ── Combined save ──────────────────────────────────────────────────────────────

def save_output(
    found: dict,
    path: str,
    cfg: dict,
    stats: Optional[dict] = None,
) -> None:
    """
    Persist current results to disk.

    Always writes a CSV companion file (used by ``load_existing_output``
    for resume support on the next run).

    Also writes a styled XLSX workbook (with Run Stats sheet) when
    ``cfg["output_format"]`` is ``"xlsx"`` (the default).

    Parameters
    ----------
    found : Dict of contact records.
    path  : Primary output path (e.g. ``"found_contacts_20241201.xlsx"``).
    cfg   : Config dict.
    stats : Optional run statistics for the Run Stats sheet.
    """
    _save_csv(found, path, cfg)
    if cfg.get("output_format", "xlsx") == "xlsx":
        _save_xlsx(found, path, cfg, stats or {})
